import os
import time
import json
import requests
import threading
from datetime import datetime, timedelta
from flask import Flask, request, jsonify, send_file
import io

app = Flask(__name__)

# ==========================================
# تحويل الأرقام العربية للإنجليزية
# ==========================================
def normalize(text):
    arabic = '٠١٢٣٤٥٦٧٨٩'
    for i, a in enumerate(arabic):
        text = text.replace(a, str(i))
    return text.strip()

# ==========================================
# الإعدادات
# ==========================================
INSTANCE_ID       = os.environ.get("INSTANCE_ID", "7107579979")
API_TOKEN         = os.environ.get("API_TOKEN", "5c1dd144d2ff4079b484b1362e763bc18dc5ebfc12e049acbe")
BASE_URL          = f"https://7107.api.greenapi.com/waInstance{INSTANCE_ID}"
BANK_ACCOUNT      = "SA6778000000001294485647"
EXPORT_SECRET     = os.environ.get("EXPORT_SECRET", "ms-export-2026")
ADMIN_GROUP       = "120363406973437339@g.us"
CONTROL_GROUP     = "120363425363360676@g.us"
SUBSCRIBERS_GROUP = "120363406971255280@g.us"
ADMIN_PHONES      = {"966531157747"}

# ==========================================
# الخدمات
# ==========================================
SERVICES = {
    "1": "الخدمات الهندسية",
    "2": "الخدمات العقارية",
    "3": "الخدمات الطلابية",
    "4": "مناديب التوصيل",
    "5": "شاليهات",
    "6": "صهريج مياه",
    "7": "اسطوانات الغاز",
    "8": "سطحات",
}

# المدينة الحالية
CITIES = {"1": "حائل"}

# ==========================================
# نظام الترجمة — 3 لغات
# ==========================================
T = {
    "ar": {
        "welcome": (
            'مرحباً بك في منصة "مذكرة سلمان" 📒\n'
            'منصتك الذكية للتواصل مع مقدمي الخدمات في مدينتك\n'
            'لاختيار الخدمة حدد مدينتك\n'
            '1 - حائل 📍\n'
            '_(مدن أخرى قريباً 🔜)_\n'
            '2 - الإدارة\n'
            '3 - اختر لغتك / Change language / زبان تبدیل کریں'
        ),
        "choose_service": (
            "اختر الخدمة:\n\n"
            "1 - الخدمات الهندسية\n"
            "2 - الخدمات العقارية\n"
            "3 - الخدمات الطلابية\n"
            "4 - مناديب التوصيل\n"
            "5 - شاليهات\n"
            "6 - صهريج مياه\n"
            "7 - اسطوانات الغاز\n"
            "8 - سطحات\n\n"
            "0 - رجوع ↩️"
        ),
        "choose_language": "اختر لغتك:\n1 - العربية 🇸🇦\n2 - English 🇬🇧\n3 - اردو 🇵🇰",
        "reg_city": "اختر مدينتك للتسجيل:\n1 - حائل 📍\n(مدن أخرى قريباً 🔜)\n\n0 - رجوع ↩️",
        "reg_service": (
            "اختر تخصصك:\n\n"
            "1 - الخدمات الهندسية\n"
            "2 - الخدمات العقارية\n"
            "3 - الخدمات الطلابية\n"
            "4 - مناديب التوصيل\n"
            "5 - شاليهات\n"
            "6 - صهريج مياه\n"
            "7 - اسطوانات الغاز\n"
            "8 - سطحات\n\n"
            "0 - رجوع ↩️"
        ),
        "reg_info": "أرسل اسمك أو اسم نشاطك التجاري\nمع رقم هويتك أو سجلك التجاري\n\n0 - رجوع ↩️",
        "reg_pending": (
            "أهلاً بك في مذكرة سلمان 🌟\n\n"
            "شكراً لاهتمامك بالانضمام إلى منصتنا\n"
            "سنقوم بمراجعة بياناتك والتحقق منها\n"
            "وفي حال كانت سليمة سيتم إشعارك\n"
            "بتفعيل حسابك كمقدم خدمة ✅"
        ),
        "reg_approved": (
            "تم اعتمادك في منصة مذكرة سلمان 🎉\n"
            "اشتراكك بدأ من اليوم لمدة 28 يوم\n"
            "ستصلك طلبات العملاء مباشرة\n"
            "أرسل 1 لاستلام أي طلب ✅"
        ),
        "admin_menu": "الإدارة:\n\n1 - تسجيل نشاطك التجاري\n2 - شكوى أو اقتراح\n3 - تواصل مع الإدارة\n0 - رجوع ↩️",
        "complaint_prompt": "اكتب شكواك أو اقتراحك وسيتم مراجعته فوراً:\n\n0 - رجوع ↩️",
        "complaint_done": "تم استلام شكواك ✅\nسيتم التواصل معك قريباً",
        "terms_client": (
            "قبل المتابعة يرجى قراءة الشروط:\n\n"
            "1️⃣ المنصة وسيط إلكتروني فقط\n"
            "ولا تتحمل مسؤولية جودة الخدمة\n\n"
            "2️⃣ يُمنع استخدام ألفاظ مسيئة\n"
            "أو التحرش أو التهديد\n\n"
            "3️⃣ في حال الإساءة يحق للمنصة\n"
            "إيقاف حسابك وإحالتك للجهات القانونية\n\n"
            "4️⃣ بياناتك الشخصية محفوظة\n"
            "ولن تُشارك مع أي طرف ثالث\n\n"
            "هل توافق؟\n\n"
            "1 - أوافق ✅\n"
            "2 - لا أوافق ❌\n"
            "0 - رجوع ↩️"
        ),
        "invalid": "الرجاء ارسال رقم صحيح",
        "desc_prompt": "اكتب وصفاً قصيراً عن طلبك:\n\n0 - رجوع ↩️",
        "chosen": "اخترت: {service}\n\n",
        "no_provider": "لم نجد لك مقدم خدمة حتى الآن 😔\n\n1 - انتظار ⏳\n2 - إلغاء الطلب ❌",
        "order_received": "تم استلام طلبك ✅\nرقم الطلب: {oid}\n\nسيتم التواصل معك خلال 5 دقائق",
        "order_accepted": "ابشر به 🎉\n\nتم قبول طلبك رقم {oid}\nالمدينة: {city}\nالخدمة: {service}\n\nمقدم الخدمة: {name}\nللتواصل: {phone}",
        "rating": "كيف كانت تجربتك مع مقدم الخدمة؟\n\n1 - ممتاز تم الاتفاق ✅\n2 - لم يتم الاتفاق\n3 - تواصل مع الإدارة",
        "reason": "ما سبب عدم الاتفاق؟\n\n1 - السعر مرتفع\n2 - لم يتجاوب\n3 - سبب آخر\n0 - رجوع ↩️",
    },
    "en": {
        "welcome": (
            'Welcome to "Mudhakkira Salman" platform 📒\n'
            'Your smart platform to connect with service providers in your city\n'
            'Choose your city to select a service\n'
            '1 - Hail 📍\n'
            '_(More cities coming soon 🔜)_\n'
            '2 - Administration\n'
            '3 - اختر لغتك / Change language / زبان تبدیل کریں'
        ),
        "choose_service": (
            "Choose a service:\n\n"
            "1 - Engineering\n"
            "2 - Real Estate\n"
            "3 - Academic\n"
            "4 - Delivery\n"
            "5 - Chalets\n"
            "6 - Water Tanker\n"
            "7 - Gas Cylinders\n"
            "8 - Tow Truck\n\n"
            "0 - Back ↩️"
        ),
        "choose_language": "Choose your language:\n1 - العربية 🇸🇦\n2 - English 🇬🇧\n3 - اردو 🇵🇰",
        "reg_city": "Choose your city to register:\n1 - Hail 📍\n(More cities coming soon 🔜)\n\n0 - Back ↩️",
        "reg_service": (
            "Choose your specialty:\n\n"
            "1 - Engineering\n"
            "2 - Real Estate\n"
            "3 - Academic\n"
            "4 - Delivery\n"
            "5 - Chalets\n"
            "6 - Water Tanker\n"
            "7 - Gas Cylinders\n"
            "8 - Tow Truck\n\n"
            "0 - Back ↩️"
        ),
        "reg_info": "Send your name or business name\nwith your ID number or commercial registration\n\n0 - Back ↩️",
        "reg_pending": (
            "Welcome to Mudhakkira Salman 🌟\n\n"
            "Thank you for your interest in joining our platform\n"
            "We will review your information and verify it\n"
            "If everything checks out, you will be notified\n"
            "when your account is activated as a service provider ✅"
        ),
        "reg_approved": (
            "You have been approved on Mudhakkira Salman 🎉\n"
            "Your subscription started today for 28 days\n"
            "You will receive client requests directly\n"
            "Send 1 to accept any request ✅"
        ),
        "admin_menu": "Administration:\n\n1 - Register your business\n2 - Complaint or suggestion\n3 - Contact admin\n0 - Back ↩️",
        "complaint_prompt": "Write your complaint or suggestion:\n\n0 - Back ↩️",
        "complaint_done": "Your complaint has been received ✅\nWe will contact you soon",
        "terms_client": (
            "Please read the terms before continuing:\n\n"
            "1️⃣ The platform is an electronic intermediary only\n\n"
            "2️⃣ No offensive language or harassment allowed\n\n"
            "3️⃣ Personal data is protected\n\n"
            "Do you agree?\n\n"
            "1 - I agree ✅\n"
            "2 - I disagree ❌\n"
            "0 - Back ↩️"
        ),
        "invalid": "Please send a valid number",
        "desc_prompt": "Write a brief description of your request:\n\n0 - Back ↩️",
        "chosen": "You chose: {service}\n\n",
        "no_provider": "No service provider found yet 😔\n\n1 - Wait ⏳\n2 - Cancel request ❌",
        "order_received": "Your request has been received ✅\nOrder number: {oid}\n\nWe will contact you within 5 minutes",
        "order_accepted": "Great news 🎉\n\nYour order {oid} has been accepted\nCity: {city}\nService: {service}\n\nProvider: {name}\nContact: {phone}",
        "rating": "How was your experience?\n\n1 - Excellent, deal done ✅\n2 - No deal\n3 - Contact support",
        "reason": "Why no deal?\n\n1 - Price too high\n2 - No response\n3 - Other\n0 - Back ↩️",
    },
    "ur": {
        "welcome": (
            '"مذکرہ سلمان" پلیٹ فارم میں خوش آمدید 📒\n'
            'آپ کے شہر میں سروس فراہم کنندگان سے جڑنے کا ذہین پلیٹ فارم\n'
            'سروس چننے کے لیے اپنا شہر منتخب کریں\n'
            '1 - حائل 📍\n'
            '_(مزید شہر جلد 🔜)_\n'
            '2 - انتظامیہ\n'
            '3 - اختر لغتك / Change language / زبان تبدیل کریں'
        ),
        "choose_service": (
            "سروس چنیں:\n\n"
            "1 - انجینئرنگ\n"
            "2 - رئیل اسٹیٹ\n"
            "3 - تعلیمی\n"
            "4 - ڈیلیوری\n"
            "5 - شالیہات\n"
            "6 - واٹر ٹینکر\n"
            "7 - گیس سلنڈر\n"
            "8 - ٹو ٹرک\n\n"
            "0 - واپس ↩️"
        ),
        "choose_language": "اپنی زبان چنیں:\n1 - العربية 🇸🇦\n2 - English 🇬🇧\n3 - اردو 🇵🇰",
        "reg_city": "رجسٹریشن کے لیے شہر چنیں:\n1 - حائل 📍\n(مزید شہر جلد 🔜)\n\n0 - واپس ↩️",
        "reg_service": (
            "اپنی خصوصیت چنیں:\n\n"
            "1 - انجینئرنگ\n"
            "2 - رئیل اسٹیٹ\n"
            "3 - تعلیمی\n"
            "4 - ڈیلیوری\n"
            "5 - شالیہات\n"
            "6 - واٹر ٹینکر\n"
            "7 - گیس سلنڈر\n"
            "8 - ٹو ٹرک\n\n"
            "0 - واپس ↩️"
        ),
        "reg_info": "اپنا نام یا کاروباری نام بھیجیں\nشناختی کارڈ یا تجارتی رجسٹریشن نمبر کے ساتھ\n\n0 - واپس ↩️",
        "reg_pending": (
            "مذکرہ سلمان میں خوش آمدید 🌟\n\n"
            "ہمارے پلیٹ فارم میں شامل ہونے میں دلچسپی کا شکریہ\n"
            "ہم آپ کی معلومات کا جائزہ لیں گے اور تصدیق کریں گے\n"
            "اگر سب کچھ درست ہوا تو آپ کو اطلاع دی جائے گی\n"
            "جب آپ کا اکاؤنٹ سروس فراہم کنندہ کے طور پر فعال ہو جائے گا ✅"
        ),
        "reg_approved": (
            "مذکرہ سلمان پر آپ کی منظوری ہو گئی 🎉\n"
            "آپ کی سبسکرپشن آج سے 28 دن کے لیے شروع ہوئی\n"
            "کلائنٹ کی درخواستیں براہ راست آپ کو ملیں گی\n"
            "کوئی بھی درخواست قبول کرنے کے لیے 1 بھیجیں ✅"
        ),
        "admin_menu": "انتظامیہ:\n\n1 - اپنا کاروبار رجسٹر کریں\n2 - شکایت یا تجویز\n3 - انتظامیہ سے رابطہ\n0 - واپس ↩️",
        "complaint_prompt": "اپنی شکایت یا تجویز لکھیں:\n\n0 - واپس ↩️",
        "complaint_done": "آپ کی شکایت موصول ہو گئی ✅\nہم جلد رابطہ کریں گے",
        "terms_client": (
            "جاری رکھنے سے پہلے شرائط پڑھیں:\n\n"
            "1️⃣ پلیٹ فارم صرف الیکٹرونک وسیط ہے\n\n"
            "2️⃣ توہین آمیز زبان ممنوع ہے\n\n"
            "3️⃣ ذاتی ڈیٹا محفوظ ہے\n\n"
            "کیا آپ متفق ہیں؟\n\n"
            "1 - متفق ✅\n"
            "2 - غیر متفق ❌\n"
            "0 - واپس ↩️"
        ),
        "invalid": "براہ کرم درست نمبر بھیجیں",
        "desc_prompt": "اپنی درخواست کی مختصر تفصیل لکھیں:\n\n0 - واپس ↩️",
        "chosen": "آپ نے چنا: {service}\n\n",
        "no_provider": "ابھی تک کوئی سروس فراہم کنندہ نہیں ملا 😔\n\n1 - انتظار ⏳\n2 - درخواست منسوخ ❌",
        "order_received": "آپ کی درخواست موصول ہو گئی ✅\nآرڈر نمبر: {oid}\n\n5 منٹ میں رابطہ کیا جائے گا",
        "order_accepted": "خوشخبری 🎉\n\nآپ کا آرڈر {oid} قبول ہو گیا\nشہر: {city}\nسروس: {service}\n\nفراہم کنندہ: {name}\nرابطہ: {phone}",
        "rating": "سروس فراہم کنندہ کے ساتھ تجربہ کیسا رہا؟\n\n1 - بہترین ✅\n2 - معاہدہ نہیں ہوا\n3 - انتظامیہ سے رابطہ",
        "reason": "معاہدہ کیوں نہیں ہوا؟\n\n1 - قیمت زیادہ\n2 - جواب نہیں دیا\n3 - اور وجہ\n0 - واپس ↩️",
    },
}

def t(phone, key, **kwargs):
    lang = user_language.get(phone, "ar")
    text = T.get(lang, T["ar"]).get(key, T["ar"].get(key, ""))
    if kwargs:
        text = text.format(**kwargs)
    return text

# ==========================================
# Render Disk
# ==========================================
DATA_PATH         = "/opt/render/project/data"
PROVIDERS_FILE    = f"{DATA_PATH}/providers.json"
CLIENTS_FILE      = f"{DATA_PATH}/clients.json"
ORDERS_FILE       = f"{DATA_PATH}/orders.json"
COUNTER_FILE      = f"{DATA_PATH}/counter.json"
LOG_FILE          = f"{DATA_PATH}/activity_log.json"
LANGUAGES_FILE    = f"{DATA_PATH}/languages.json"
PENDING_FILE      = f"{DATA_PATH}/pending_approval.json"

# ==========================================
# البيانات في الذاكرة
# ==========================================
user_sessions     = {}
control_sessions  = {}
user_language     = {}
pending_approval  = {}
provider_last_order = {}  # آخر طلب أُرسل لكل مقدم {phone: oid}
registered_clients   = set()
registered_providers = {}
pending_orders    = {}
blocked_users     = {}
activity_log      = []
order_counter     = [1000]
last_activity     = {}
SESSION_TIMEOUT   = 120  # دقيقتان

# ==========================================
# حفظ وتحميل
# ==========================================
def load_data():
    global registered_providers, registered_clients, pending_orders, order_counter
    try:
        os.makedirs(DATA_PATH, exist_ok=True)
        if os.path.exists(PROVIDERS_FILE):
            with open(PROVIDERS_FILE, "r", encoding="utf-8") as f:
                registered_providers = json.load(f)
            print(f"✅ مقدمون: {len(registered_providers)}")
        if os.path.exists(CLIENTS_FILE):
            with open(CLIENTS_FILE, "r", encoding="utf-8") as f:
                registered_clients = set(json.load(f))
            print(f"✅ عملاء: {len(registered_clients)}")
        if os.path.exists(COUNTER_FILE):
            with open(COUNTER_FILE, "r", encoding="utf-8") as f:
                order_counter[0] = json.load(f).get("counter", 1000)
        if os.path.exists(LOG_FILE):
            with open(LOG_FILE, "r", encoding="utf-8") as f:
                activity_log.extend(json.load(f))
        if os.path.exists(LANGUAGES_FILE):
            with open(LANGUAGES_FILE, "r", encoding="utf-8") as f:
                user_language.update(json.load(f))
        if os.path.exists(PENDING_FILE):
            with open(PENDING_FILE, "r", encoding="utf-8") as f:
                pending_approval.update(json.load(f))
        if os.path.exists(ORDERS_FILE):
            with open(ORDERS_FILE, "r", encoding="utf-8") as f:
                saved = json.load(f)
            for oid, od in saved.items():
                if not od.get("taken") and od.get("providers"):
                    od["taken"] = False
                    pending_orders[oid] = od
                    threading.Timer(2, broadcast_order, args=[oid]).start()
            print(f"✅ طلبات معلقة: {len(pending_orders)}")
    except Exception as e:
        print(f"خطأ تحميل: {e}")

def save_providers():
    try:
        with open(PROVIDERS_FILE, "w", encoding="utf-8") as f:
            json.dump(registered_providers, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"خطأ: {e}")

def save_clients():
    try:
        with open(CLIENTS_FILE, "w", encoding="utf-8") as f:
            json.dump(list(registered_clients), f, ensure_ascii=False)
    except Exception as e:
        print(f"خطأ: {e}")

def save_counter():
    try:
        with open(COUNTER_FILE, "w", encoding="utf-8") as f:
            json.dump({"counter": order_counter[0]}, f)
    except Exception as e:
        print(f"خطأ: {e}")

def save_orders():
    try:
        with open(ORDERS_FILE, "w", encoding="utf-8") as f:
            json.dump(pending_orders, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"خطأ: {e}")

def save_language():
    try:
        with open(LANGUAGES_FILE, "w", encoding="utf-8") as f:
            json.dump(user_language, f, ensure_ascii=False)
    except Exception as e:
        print(f"خطأ: {e}")

def save_pending():
    try:
        with open(PENDING_FILE, "w", encoding="utf-8") as f:
            json.dump(pending_approval, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"خطأ: {e}")

def log_event(event_type, phone, details="", order_id=""):
    entry = {
        "time":     datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "type":     event_type,
        "phone":    phone,
        "order_id": order_id,
        "details":  details,
    }
    activity_log.append(entry)
    try:
        with open(LOG_FILE, "w", encoding="utf-8") as f:
            json.dump(activity_log[-5000:], f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"خطأ سجل: {e}")

# ==========================================
# دوال الإرسال
# ==========================================
def send_msg(to, text):
    chat_id = f"{to}@c.us" if "@" not in to else to
    try:
        r = requests.post(
            f"{BASE_URL}/sendMessage/{API_TOKEN}",
            json={"chatId": chat_id, "message": text},
            timeout=10
        )
        return r.json().get("idMessage", "")
    except Exception as e:
        print(f"Send error: {e}")
        return ""

def send_group(gid, text):
    try:
        r = requests.post(
            f"{BASE_URL}/sendMessage/{API_TOKEN}",
            json={"chatId": gid, "message": text},
            timeout=10
        )
        return r.json().get("idMessage", "")
    except Exception as e:
        print(f"Group error: {e}")
        return ""

# ==========================================
# دوال مساعدة
# ==========================================
def is_blocked(phone):
    if phone in blocked_users:
        until = blocked_users[phone]
        if time.time() < until:
            return True, int((until - time.time()) / 60)
        del blocked_users[phone]
    return False, 0

def check_timeout(phone):
    now = time.time()
    if phone in last_activity:
        if now - last_activity[phone] > SESSION_TIMEOUT:
            user_sessions.pop(phone, None)
            last_activity[phone] = now
            return True
    last_activity[phone] = now
    return False

def check_subscription(provider):
    expiry = provider.get("expiry", "")
    if not expiry:
        return True
    try:
        return datetime.now() < datetime.strptime(expiry, "%Y-%m-%d")
    except:
        return True

def count_providers(city, service):
    return sum(
        1 for d in registered_providers.values()
        if d.get("city") == city
        and d.get("specialty") == service
        and d.get("status") == "active"
    )

# ==========================================
# منطق الطلبات
# ==========================================
def create_order(phone, city, service, description=""):
    order_counter[0] += 1
    oid = f"MS-{order_counter[0]}"
    matched = [
        p for p, d in registered_providers.items()
        if d.get("city") == city
        and d.get("specialty") == service
        and d.get("status") == "active"
        and check_subscription(d)
    ]
    pending_orders[oid] = {
        "phone":       phone,
        "city":        city,
        "service":     service,
        "description": description,
        "attempts":    1,
        "blocked":     [],
        "taken":       False,
        "providers":   matched,
        "created":     datetime.now().strftime("%Y-%m-%d %H:%M"),
    }
    save_counter()
    save_orders()
    log_event("طلب_جديد", phone, f"{city} | {service} | {description}", oid)
    user_sessions[phone] = {"step": "waiting", "order_id": oid}
    send_msg(phone, t(phone, "order_received", oid=oid))
    if not matched:
        send_group(ADMIN_GROUP,
            f"⚠️ لا يوجد مقدم\nالطلب: {oid}\nالمدينة: {city}\nالخدمة: {service}\nالعميل: {phone}"
        )
    def start_broadcast():
        time.sleep(1)
        broadcast_order(oid)
    threading.Thread(target=start_broadcast).start()

def broadcast_order(oid):
    if oid not in pending_orders:
        return
    od        = pending_orders[oid]
    providers = od.get("providers", [])
    blocked   = od.get("blocked", [])
    cp        = od["phone"]
    available = [p for p in providers if p not in blocked]
    desc      = f"الوصف: {od['description']}\n" if od.get("description") else ""
    for p in available:
        send_msg(p,
            f"طلب جديد 🔔\n"
            f"رقم الطلب: {oid}\n"
            f"المدينة: {od['city']}\n"
            f"الخدمة: {od['service']}\n"
            f"{desc}"
            f"━━━━━━━━━━━━━━\n"
            f"📌 لاستلام هذا الطلب:\n"
            f"اضغط مطولاً على هذه الرسالة\n"
            f"ثم اختر (رد) واكتب: 1"
        )
        provider_last_order[p] = oid
        time.sleep(0.3)
    def check_5min(cp=cp, oid=oid):
        time.sleep(5 * 60)
        if oid not in pending_orders or pending_orders.get(oid, {}).get("taken"):
            return
        user_sessions[cp] = {"step": "waiting_choice", "order_id": oid}
        send_msg(cp, t(cp, "no_provider"))
    threading.Thread(target=check_5min).start()

def extract_oid_from_quoted(md):
    """استخراج رقم الطلب من نص الرسالة المردود عليها"""
    try:
        ext = md.get("extendedTextMessageData", {})
        # Green API يضع السياق في stanzaId أو quotedMessage
        ctx = ext.get("contextInfo", {}) or ext.get("quotedMessageData", {})
        quoted_text = (
            ctx.get("quotedMessage", {}).get("conversation", "")
            or ctx.get("quotedMessage", {}).get("extendedTextMessage", {}).get("text", "")
            or ctx.get("quotedBody", "")
            or ext.get("description", "")
        )
        print(f"🔍 quoted_text: {quoted_text[:100] if quoted_text else 'فارغ'}")
        if not quoted_text:
            return None
        import re
        match = re.search(r"MS-\\d+", quoted_text)
        if match:
            return match.group(0)
    except Exception as e:
        print(f"extract error: {e}")
    return None

def handle_provider_accept(phone, quoted_oid=None):
    # أولوية: الطلب المستخرج من الرد → آخر طلب أُرسل → أي طلب متاح
    candidates = []
    if quoted_oid and quoted_oid in pending_orders:
        candidates.append(quoted_oid)
    last = provider_last_order.get(phone)
    if last and last in pending_orders and last not in candidates:
        candidates.append(last)
    for oid in pending_orders:
        if oid not in candidates:
            candidates.append(oid)

    for oid in candidates:
        od = pending_orders.get(oid)
        if not od:
            continue
        if od.get("taken"):
            continue
        if phone not in od.get("providers", []):
            continue
        if phone in od.get("blocked", []):
            continue
        cp   = od["phone"]
        name = registered_providers.get(phone, {}).get("name", "مقدم الخدمة")
        od["taken"] = True
        od["blocked"].append(phone)
        provider_last_order.pop(phone, None)
        log_event("قبول_طلب", phone, f"عميل: {cp}", oid)
        send_msg(cp, t(cp, "order_accepted",
            oid=oid, city=od["city"], service=od["service"], name=name, phone=phone))
        send_msg(phone, f"تم تأكيد استلامك للطلب {oid} ✅\nسيتواصل معك العميل قريباً")
        def send_rating(cp=cp):
            time.sleep(60)
            send_msg(cp, t(cp, "rating"))
        threading.Thread(target=send_rating).start()
        user_sessions[cp] = {"step": "provider_sent", "order_id": oid}
        save_orders()
        return
    send_msg(phone, "لا يوجد طلب متاح لك الآن ⏳")

def resend_order(phone, oid, reason, price=None):
    if oid not in pending_orders:
        return
    od = pending_orders[oid]
    od["attempts"] += 1
    od["taken"] = False
    if price:
        od["last_price"] = price
    attempts = od["attempts"]
    warning  = "\nتنبيه: هذه آخر محاولة ⚠️" if attempts == 3 else ""
    send_msg(phone, f"تم إعادة طلبك\nالمحاولة {attempts} من 3{warning}\nسيتم التواصل معك خلال 5 دقائق")
    user_sessions[phone] = {"step": "waiting", "order_id": oid}
    save_orders()
    broadcast_order(oid)

# ==========================================
# معالج العميل
# ==========================================
def handle_customer(phone, msg):
    blocked, remaining = is_blocked(phone)
    if blocked:
        send_msg(phone, f"حسابك موقوف مؤقتاً\nالمتبقي: {remaining} دقيقة ⏱️")
        return
    timed_out = check_timeout(phone)
    if timed_out:
        user_sessions[phone] = {"step": "start"}
    session = user_sessions.get(phone, {"step": "start"})
    step    = session.get("step", "start")

    # ── البداية ──
    if step == "start":
        log_event("رسالة_جديدة", phone, "بدأ المحادثة")
        send_msg(phone, t(phone, "welcome"))
        user_sessions[phone] = {"step": "city"}

    # ── اختيار اللغة ──
    elif step == "choose_language":
        lang_map = {"1": "ar", "2": "en", "3": "ur"}
        if msg not in lang_map:
            send_msg(phone, T["ar"]["choose_language"])
            return
        user_language[phone] = lang_map[msg]
        save_language()
        send_msg(phone, t(phone, "welcome"))
        user_sessions[phone] = {"step": "city"}

    # ── القائمة الرئيسية ──
    elif step == "city":
        if msg == "1":
            log_event("اختيار_مدينة", phone, "حائل")
            user_sessions[phone] = {"step": "service", "city": "حائل"}
            send_msg(phone, t(phone, "choose_service"))
        elif msg == "2":
            user_sessions[phone] = {"step": "admin_menu"}
            send_msg(phone, t(phone, "admin_menu"))
        elif msg == "3":
            user_sessions[phone] = {"step": "choose_language"}
            send_msg(phone, T["ar"]["choose_language"])
        else:
            send_msg(phone, t(phone, "invalid"))

    # ── اختيار الخدمة ──
    elif step == "service":
        city = session.get("city", "حائل")
        if msg == "0":
            user_sessions[phone] = {"step": "city"}
            send_msg(phone, t(phone, "welcome"))
            return
        if msg not in SERVICES:
            send_msg(phone, t(phone, "invalid"))
            return
        service = SERVICES[msg]
        log_event("اختيار_خدمة", phone, f"{city} | {service}")
        user_sessions[phone] = {"step": "description", "city": city, "service": service}
        send_msg(phone, t(phone, "chosen", service=service) + t(phone, "desc_prompt"))

    # ── وصف الطلب ──
    elif step == "description":
        city    = session.get("city", "حائل")
        service = session.get("service", "")
        if msg == "0":
            user_sessions[phone] = {"step": "service", "city": city}
            send_msg(phone, t(phone, "choose_service"))
            return
        if phone not in registered_clients:
            user_sessions[phone] = {"step": "terms", "city": city, "service": service, "description": msg}
            send_msg(phone, t(phone, "terms_client"))
        else:
            create_order(phone, city, service, msg)

    # ── شروط العميل ──
    elif step == "terms":
        city        = session.get("city", "حائل")
        service     = session.get("service", "")
        description = session.get("description", "")
        if msg == "0":
            user_sessions[phone] = {"step": "description", "city": city, "service": service}
            send_msg(phone, t(phone, "chosen", service=service) + t(phone, "desc_prompt"))
            return
        if msg == "1":
            registered_clients.add(phone)
            save_clients()
            create_order(phone, city, service, description)
        elif msg == "2":
            send_msg(phone, "شكراً لك 🌟")
            user_sessions[phone] = {"step": "start"}
        else:
            send_msg(phone, t(phone, "terms_client"))

    # ── الانتظار ──
    elif step == "waiting":
        return  # صمت تام

    # ── انتظار أو إلغاء ──
    elif step == "waiting_choice":
        oid = session.get("order_id")
        if msg == "1":
            user_sessions[phone] = {"step": "waiting", "order_id": oid}
            send_msg(phone, "شكراً لصبرك ⏳\nسنواصل البحث")
            if oid in pending_orders:
                pending_orders[oid]["blocked"] = []
                broadcast_order(oid)
        elif msg == "2":
            log_event("إلغاء_طلب", phone, "", oid)
            pending_orders.pop(oid, None)
            save_orders()
            user_sessions[phone] = {"step": "start"}
            send_msg(phone, "تم إلغاء طلبك ✅\nيمكنك إرسال طلب جديد في أي وقت")
        else:
            send_msg(phone, t(phone, "no_provider"))

    # ── التقييم ──
    elif step == "provider_sent":
        oid = session.get("order_id")
        od  = pending_orders.get(oid, {})
        if msg == "1":
            log_event("اتفاق_ناجح", phone, "", oid)
            pending_orders.pop(oid, None)
            save_orders()
            user_sessions[phone] = {"step": "start"}
            send_msg(phone, "ممتاز! نتمنى لك تجربة رائعة مع مذكرة سلمان 🌟")
        elif msg == "2":
            attempts = od.get("attempts", 1)
            if attempts >= 3:
                blocked_users[phone] = time.time() + 15 * 60
                log_event("حظر_مؤقت", phone, "3 محاولات فاشلة", oid)
                pending_orders.pop(oid, None)
                save_orders()
                user_sessions[phone] = {"step": "start"}
                send_msg(phone, "تم استنفاد المحاولات\nحسابك موقوف 15 دقيقة ⏱️")
            else:
                user_sessions[phone] = {"step": "reason", "order_id": oid}
                send_msg(phone, t(phone, "reason"))
        elif msg == "3":
            send_group(ADMIN_GROUP, f"🚨 شكوى عميل\nالطلب: {oid}\nالعميل: {phone}")
            send_msg(phone, "سيتواصل معك فريق الإدارة قريباً 🙏")
            user_sessions[phone] = {"step": "start"}
        else:
            send_msg(phone, t(phone, "rating"))

    # ── سبب عدم الاتفاق ──
    elif step == "reason":
        oid = session.get("order_id")
        if msg == "0":
            user_sessions[phone] = {"step": "provider_sent", "order_id": oid}
            send_msg(phone, t(phone, "rating"))
            return
        if msg == "1":
            user_sessions[phone] = {"step": "price", "order_id": oid}
            send_msg(phone, "كم السعر المعروض؟ (بالريال)")
        elif msg == "2":
            resend_order(phone, oid, "لم يتجاوب")
        elif msg == "3":
            user_sessions[phone] = {"step": "custom_reason", "order_id": oid}
            send_msg(phone, "اكتب سبب عدم الاتفاق:")
        else:
            send_msg(phone, t(phone, "reason"))

    elif step == "price":
        resend_order(phone, session.get("order_id"), "السعر مرتفع", price=msg)

    elif step == "custom_reason":
        resend_order(phone, session.get("order_id"), msg)

    # ── الشكاوى ──
    elif step == "complaint":
        if msg == "0":
            user_sessions[phone] = {"step": "admin_menu"}
            send_msg(phone, t(phone, "admin_menu"))
            return
        log_event("شكوى", phone, msg)
        send_group(ADMIN_GROUP, f"🚨 شكوى\nالعميل: {phone}\nالشكوى: {msg}")
        send_msg(phone, t(phone, "complaint_done"))
        user_sessions[phone] = {"step": "start"}

    # ── قائمة الإدارة ──
    elif step == "admin_menu":
        if msg == "0":
            user_sessions[phone] = {"step": "city"}
            send_msg(phone, t(phone, "welcome"))
            return
        if msg == "1":
            user_sessions[phone] = {"step": "reg_city"}
            send_msg(phone, t(phone, "reg_city"))
        elif msg == "2":
            user_sessions[phone] = {"step": "complaint"}
            send_msg(phone, t(phone, "complaint_prompt"))
        elif msg == "3":
            send_msg(phone, "سيتواصل معك فريق الإدارة قريباً 🙏")
            send_group(ADMIN_GROUP, f"📞 طلب تواصل\nرقم: {phone}")
            user_sessions[phone] = {"step": "start"}
        else:
            send_msg(phone, t(phone, "admin_menu"))

    # ── تسجيل مقدم — المدينة ──
    elif step == "reg_city":
        if msg == "0":
            user_sessions[phone] = {"step": "city"}
            send_msg(phone, t(phone, "welcome"))
            return
        if msg != "1":
            send_msg(phone, t(phone, "reg_city"))
            return
        user_sessions[phone] = {"step": "reg_service", "reg_city": "حائل"}
        send_msg(phone, t(phone, "reg_service"))

    # ── تسجيل مقدم — التخصص ──
    elif step == "reg_service":
        if msg == "0":
            user_sessions[phone] = {"step": "reg_city"}
            send_msg(phone, t(phone, "reg_city"))
            return
        if msg not in SERVICES:
            send_msg(phone, t(phone, "invalid"))
            return
        user_sessions[phone].update({"step": "reg_info", "reg_service": SERVICES[msg]})
        send_msg(phone, t(phone, "reg_info"))

    # ── تسجيل مقدم — البيانات ──
    elif step == "reg_info":
        if msg == "0":
            user_sessions[phone] = {"step": "reg_service", "reg_city": session.get("reg_city", "حائل")}
            send_msg(phone, t(phone, "reg_service"))
            return
        reg_city    = session.get("reg_city", "حائل")
        reg_service = session.get("reg_service", "")
        pending_approval[phone] = {
            "name":      msg,
            "city":      reg_city,
            "service":   reg_service,
            "phone":     phone,
            "timestamp": time.time(),
        }
        save_pending()
        send_group(SUBSCRIBERS_GROUP,
            f"🆕 طلب تسجيل جديد\n"
            f"الاسم/النشاط: {msg}\n"
            f"المدينة: {reg_city}\n"
            f"التخصص: {reg_service}\n"
            f"الرقم: {phone}\n"
            f"تفاعل مع هذه الرسالة لاعتماده ✅"
        )
        send_msg(phone, t(phone, "reg_pending"))
        user_sessions[phone] = {"step": "reg_pending"}
        log_event("طلب_تسجيل", phone, f"{reg_city} | {reg_service} | {msg}")

    # ── صمت انتظار الاعتماد ──
    elif step == "reg_pending":
        return  # صمت تام

# ==========================================
# قائمة مقدم الخدمة المسجّل
# ==========================================
def handle_provider_menu(phone, msg, provider):
    session = user_sessions.get(phone, {"step": "provider_main"})
    step    = session.get("step", "provider_main")

    if step == "provider_main":
        if msg == "1":
            # حسابي
            expiry = provider.get("expiry", "")
            status = "مفعّل ✅" if provider.get("status") == "active" else "موقوف ⚠️"
            if expiry:
                try:
                    days_left = (datetime.strptime(expiry, "%Y-%m-%d") - datetime.now()).days
                    if days_left > 0:
                        sub_text = f"⏳ متبقي لك {days_left} يوم"
                    elif days_left == 0:
                        sub_text = "⚠️ اشتراكك ينتهي اليوم"
                    else:
                        sub_text = f"🚫 انتهى الاشتراك منذ {abs(days_left)} يوم"
                except:
                    sub_text = f"الاشتراك: {expiry}"
            else:
                sub_text = "الاشتراك: غير محدد"
            send_msg(phone,
                f"معلومات حسابك:\n\n"
                f"الاسم: {provider.get('name', '')}\n"
                f"المدينة: {provider.get('city', '')}\n"
                f"التخصص: {provider.get('specialty', '')}\n"
                f"الحالة: {status}\n"
                f"{sub_text}\n\n"
                f"1 - تجديد الاشتراك 🔄\n"
                f"0 - رجوع ↩️"
            )
            user_sessions[phone] = {"step": "provider_account"}
        elif msg == "2":
            user_sessions[phone] = {"step": "provider_contact"}
            send_msg(phone, "اكتب رسالتك للإدارة:\n\n0 - رجوع ↩️")
        else:
            send_msg(phone,
                f"مرحباً {provider.get('name', '')} 👋\n\n"
                "1 - حسابي 👤\n"
                "2 - تواصل مع الإدارة\n"
            )

    elif step == "provider_account":
        if msg == "1":
            # تجديد الاشتراك
            send_msg(phone,
                f"تجديد الاشتراك 🔄\n\n"
                f"💰 20 ريال لكل 28 يوم\n\n"
                f"رقم الحساب:\n{BANK_ACCOUNT}\n"
                f"أرسل إيصال التحويل للإدارة\n"
                f"وسيتم تجديد اشتراكك خلال 24 ساعة 🙏"
            )
            send_group(ADMIN_GROUP,
                f"🔄 طلب تجديد اشتراك\n"
                f"الاسم: {provider.get('name', '')}\n"
                f"الرقم: {phone}\n"
                f"التخصص: {provider.get('specialty', '')}\n"
                f"المدينة: {provider.get('city', '')}"
            )
            user_sessions[phone] = {"step": "provider_main"}
        else:
            user_sessions[phone] = {"step": "provider_main"}
            send_msg(phone,
                f"مرحباً {provider.get('name', '')} 👋\n\n"
                "1 - حسابي 👤\n"
                "2 - تواصل مع الإدارة\n"
            )

    elif step == "provider_contact":
        if msg == "0":
            user_sessions[phone] = {"step": "provider_main"}
            send_msg(phone,
                f"مرحباً {provider.get('name', '')} 👋\n\n"
                "1 - طلب جديد (كعميل)\n"
                "2 - حسابي\n"
                "3 - تواصل مع الإدارة"
            )
            return
        send_group(ADMIN_GROUP,
            f"📞 رسالة مقدم خدمة\n"
            f"الاسم: {provider.get('name', '')}\n"
            f"الرقم: {phone}\n"
            f"الرسالة: {msg}"
        )
        send_msg(phone, "تم إرسال رسالتك ✅\nسيتم التواصل معك قريباً 🙏")
        user_sessions[phone] = {"step": "provider_main"}

# ==========================================
# لوحة التحكم
# ==========================================
CTRL_MAIN = (
    "لوحة التحكم 🎮\n\n"
    "1 - رسالة جماعية للمقدمين 📢\n"
    "2 - رسالة جماعية للعملاء 👥\n"
    "3 - إدارة مقدمي الخدمة ⚙️\n"
    "4 - إدارة العملاء 👤\n"
    "5 - تحميل البيانات 📊\n"
    "0 - إلغاء ❌"
)
CTRL_PROVIDERS = (
    "إدارة مقدمي الخدمة ⚙️\n\n"
    "1 - عرض القائمة\n"
    "2 - إيقاف مقدم\n"
    "3 - تفعيل مقدم\n"
    "4 - حذف مقدم\n"
    "0 - رجوع ↩️"
)
CTRL_CLIENTS = (
    "إدارة العملاء 👤\n\n"
    "1 - عرض القائمة\n"
    "2 - حذف عميل\n"
    "3 - رفع حظر عميل\n"
    "0 - رجوع ↩️"
)
CTRL_BROADCAST = (
    "اختر المستهدفين:\n\n"
    "1 - الخدمات الهندسية\n"
    "2 - الخدمات العقارية\n"
    "3 - الخدمات الطلابية\n"
    "4 - مناديب التوصيل\n"
    "5 - شاليهات\n"
    "6 - صهريج مياه\n"
    "7 - اسطوانات الغاز\n"
    "8 - سطحات\n"
    ""
    "9  - مدينة محددة\n"
    "10 - الجميع 📢\n"
    "0  - رجوع ↩️"
)

def handle_control(phone, msg):
    session = control_sessions.get(phone, {"step": "start"})
    step    = session.get("step", "start")

    if msg == "تحكم" or step in ["start", ""]:
        control_sessions[phone] = {"step": "main"}
        send_msg(phone, CTRL_MAIN)
        return

    if msg == "0":
        control_sessions[phone] = {"step": "main"}
        send_msg(phone, CTRL_MAIN)
        return

    if step == "main":
        if msg == "1":
            control_sessions[phone] = {"step": "broadcast"}
            send_msg(phone, CTRL_BROADCAST)
        elif msg == "2":
            control_sessions[phone] = {"step": "write_clients"}
            send_msg(phone, f"عدد العملاء: {len(registered_clients)}\n\nاكتب رسالتك:\n0 - رجوع ↩️")
        elif msg == "3":
            control_sessions[phone] = {"step": "manage_providers"}
            send_msg(phone, CTRL_PROVIDERS)
        elif msg == "4":
            control_sessions[phone] = {"step": "manage_clients"}
            send_msg(phone, CTRL_CLIENTS)
        elif msg == "5":
            control_sessions[phone] = {"step": "export"}
            base = os.environ.get("RENDER_EXTERNAL_URL", "https://absherbh-bot.onrender.com")
            send_msg(phone,
                f"تحميل البيانات 📊\n\n"
                f"الرابط:\n{base}/export?key={EXPORT_SECRET}\n\n"
                f"المقدمون: {len(registered_providers)}\n"
                f"العملاء: {len(registered_clients)}\n\n"
                f"0 - رجوع ↩️"
            )
        else:
            send_msg(phone, "الرجاء ارسال رقم من 1 الى 5")

    elif step == "write_clients":
        count = 0
        for c in list(registered_clients):
            send_msg(c, msg)
            count += 1
            time.sleep(0.5)
        send_msg(phone, f"✅ تم الإرسال لـ {count} عميل")
        control_sessions[phone] = {"step": "main"}
        send_msg(phone, CTRL_MAIN)

    elif step == "manage_providers":
        if msg == "1":
            if not registered_providers:
                send_msg(phone, "لا يوجد مقدمون مسجلون")
                return
            chunk = ""
            for p, d in registered_providers.items():
                s = "✅" if d.get("status") == "active" else "⚠️"
                line = f"{s} {d.get('name','')} | {d.get('specialty','')} | {d.get('city','')} | {p}\n"
                if len(chunk) + len(line) > 3000:
                    send_msg(phone, chunk)
                    chunk = ""
                    time.sleep(0.5)
                chunk += line
            if chunk:
                send_msg(phone, chunk)
            send_msg(phone, CTRL_PROVIDERS)
        elif msg in ["2", "3", "4"]:
            action_map = {"2": "إيقاف", "3": "تفعيل", "4": "حذف"}
            control_sessions[phone] = {"step": "provider_action", "action": msg}
            send_msg(phone, f"أدخل رقم المقدم الذي تريد {action_map[msg]}ه:\n(مثال: 966501234567)\n\n0 - رجوع ↩️")
        else:
            send_msg(phone, CTRL_PROVIDERS)

    elif step == "provider_action":
        action = session.get("action")
        target = msg.strip()
        if target not in registered_providers:
            send_msg(phone, f"الرقم {target} غير موجود")
            return
        name = registered_providers[target].get("name", target)
        if action == "2":
            registered_providers[target]["status"] = "inactive"
            save_providers()
            send_msg(phone, f"✅ تم إيقاف {name}")
            send_msg(target, "تم إيقاف حسابك مؤقتاً\nللاستفسار تواصل معنا")
        elif action == "3":
            registered_providers[target]["status"] = "active"
            save_providers()
            send_msg(phone, f"✅ تم تفعيل {name}")
            send_msg(target, "تم تفعيل حسابك ✅\nستصلك الطلبات الآن")
        elif action == "4":
            del registered_providers[target]
            save_providers()
            send_msg(phone, f"✅ تم حذف {name}")
            send_msg(target, "تم حذف حسابك من المنصة")
        control_sessions[phone] = {"step": "manage_providers"}
        send_msg(phone, CTRL_PROVIDERS)

    elif step == "manage_clients":
        if msg == "1":
            if not registered_clients:
                send_msg(phone, "لا يوجد عملاء مسجلون")
                return
            chunk = ""
            for i, c in enumerate(registered_clients, 1):
                s = "🚫" if c in blocked_users and time.time() < blocked_users[c] else "✅"
                line = f"{s} {i}. {c}\n"
                if len(chunk) + len(line) > 3000:
                    send_msg(phone, chunk)
                    chunk = ""
                    time.sleep(0.5)
                chunk += line
            if chunk:
                send_msg(phone, f"عدد العملاء: {len(registered_clients)}\n\n" + chunk)
            send_msg(phone, CTRL_CLIENTS)
        elif msg in ["2", "3"]:
            action_map = {"2": "delete", "3": "unblock"}
            label_map  = {"2": "حذفه", "3": "رفع حظره"}
            control_sessions[phone] = {"step": "client_action", "action": action_map[msg]}
            send_msg(phone, f"أدخل رقم العميل الذي تريد {label_map[msg]}:\n\n0 - رجوع ↩️")
        else:
            send_msg(phone, CTRL_CLIENTS)

    elif step == "client_action":
        action = session.get("action")
        target = msg.strip()
        if action == "delete":
            if target in registered_clients:
                registered_clients.discard(target)
                save_clients()
                blocked_users.pop(target, None)
                user_sessions.pop(target, None)
                send_msg(phone, f"✅ تم حذف {target}")
                send_msg(target, "تم حذف حسابك من المنصة")
            else:
                send_msg(phone, f"الرقم {target} غير موجود")
        elif action == "unblock":
            if target in blocked_users:
                del blocked_users[target]
                send_msg(phone, f"✅ تم رفع الحظر عن {target}")
                send_msg(target, "تم رفع الحظر عن حسابك ✅")
            else:
                send_msg(phone, f"الرقم {target} غير محظور")
        control_sessions[phone] = {"step": "manage_clients"}
        send_msg(phone, CTRL_CLIENTS)

    elif step == "broadcast":
        targets = []
        label   = ""
        if msg in SERVICES:
            label   = SERVICES[msg]
            targets = [p for p, d in registered_providers.items() if d.get("specialty") == label]
        elif msg == "9":
            targets = [p for p, d in registered_providers.items() if d.get("city") == "حائل"]
            label   = "حائل"
            control_sessions[phone] = {"step": "write_providers", "targets": targets, "label": label}
            send_msg(phone, f"اخترت: {label} ({len(targets)} مقدم)\n\nاكتب رسالتك:\n0 - رجوع ↩️")
            return
        elif msg == "10":
            label   = "الجميع"
            targets = list(registered_providers.keys())
        else:
            send_msg(phone, CTRL_BROADCAST)
            return
        control_sessions[phone] = {"step": "write_providers", "targets": targets, "label": label}
        send_msg(phone, f"اخترت: {label} ({len(targets)} مقدم)\n\nاكتب رسالتك:\n0 - رجوع ↩️")

    elif step == "write_providers":
        targets = session.get("targets", [])
        label   = session.get("label", "")
        count   = 0
        for p in targets:
            send_msg(p, msg)
            count += 1
            time.sleep(0.5)
        send_msg(phone, f"✅ تم الإرسال لـ {count} مقدم في {label}")
        control_sessions[phone] = {"step": "main"}
        send_msg(phone, CTRL_MAIN)

    elif step == "export":
        control_sessions[phone] = {"step": "main"}
        send_msg(phone, CTRL_MAIN)

# ==========================================
# Webhook
# ==========================================
@app.route("/webhook", methods=["POST"])
def webhook():
    try:
        data = request.get_json()
        if not data:
            return jsonify({"status": "ok"}), 200

        wtype   = data.get("typeWebhook", "")
        sd      = data.get("senderData", {})
        md      = data.get("messageData", {})
        sender  = sd.get("sender", "")
        chat_id = sd.get("chatId", "")
        mt      = md.get("typeMessage", "")

        if wtype != "incomingMessageReceived":
            return jsonify({"status": "ok"}), 200

        # ── القروبات ──
        if "@g.us" in chat_id:
            # قروب المشتركين — أي رسالة = اعتماد المقدمين المنتظرين
            if chat_id == SUBSCRIBERS_GROUP:
                # أي نوع رسالة في القروب = اعتماد كل المقدمين المنتظرين
                approved_count = 0
                for p, d in list(pending_approval.items()):
                    if time.time() - d.get("timestamp", 0) < 24 * 3600:
                        expiry_date = (datetime.now() + timedelta(days=28)).strftime("%Y-%m-%d")
                        registered_providers[p] = {
                            "name":       d.get("name", ""),
                            "city":       d.get("city", "حائل"),
                            "specialty":  d.get("service", ""),
                            "status":     "active",
                            "expiry":     expiry_date,
                            "registered": datetime.now().strftime("%Y-%m-%d"),
                        }
                        pending_approval.pop(p)
                        user_sessions[p] = {"step": "start"}
                        send_msg(p, t(p, "reg_approved"))
                        log_event("اعتماد_مقدم", p, d.get("name", ""))
                        approved_count += 1
                        print(f"✅ اعتماد: {p}")
                if approved_count > 0:
                    save_providers()
                    save_pending()
                    print(f"✅ تم اعتماد {approved_count} مقدم")

            # قروب التحكم
            elif chat_id == CONTROL_GROUP and mt in ["textMessage", "extendedTextMessage"]:
                if mt == "textMessage":
                    text = md.get("textMessageData", {}).get("textMessage", "")
                else:
                    text = md.get("extendedTextMessageData", {}).get("text", "")
                if text:
                    text         = normalize(text)
                    sender_phone = sender.replace("@c.us", "")
                    ADMIN_PHONES.add(sender_phone)
                    ctrl = control_sessions.get(sender_phone, {"step": "start"})
                    if text == "تحكم" or ctrl.get("step") not in ["start", ""]:
                        handle_control(sender_phone, text)

            return jsonify({"status": "ok"}), 200

        # ── رسائل صوتية ──
        if mt in ["audioMessage", "pttMessage"]:
            phone = sender.replace("@c.us", "")
            send_msg(phone, "عذراً 🎤\nالرجاء إرسال رسالة نصية فقط")
            return jsonify({"status": "ok"}), 200

        # ── استخراج النص ──
        if mt == "textMessage":
            text = md.get("textMessageData", {}).get("textMessage", "")
        elif mt == "extendedTextMessage":
            text = md.get("extendedTextMessageData", {}).get("text", "")
        else:
            return jsonify({"status": "ok"}), 200

        if not text:
            return jsonify({"status": "ok"}), 200

        text  = normalize(text)
        phone = sender.replace("@c.us", "")
        print(f"📩 [{phone}] {text}")

        # ── أدمن مصرح ──
        if phone in ADMIN_PHONES:
            ctrl = control_sessions.get(phone, {"step": "start"})
            if text == "تحكم" or ctrl.get("step") not in ["start", ""]:
                handle_control(phone, text)
                return jsonify({"status": "ok"}), 200

        # ── فلتر سعودي ──
        if not phone.startswith("966"):
            send_msg(phone, "عذراً\nهذه الخدمة متاحة للأرقام السعودية فقط 🇸🇦")
            return jsonify({"status": "ok"}), 200

        # ── مقدم خدمة مسجّل ──
        if phone in registered_providers:
            session = user_sessions.get(phone, {"step": "provider_main"})
            step    = session.get("step", "provider_main")
            print(f"🔧 مقدم [{phone}] step=[{step}] text=[{text}]")
            client_steps = {
                "city", "service", "description", "terms", "waiting",
                "waiting_choice", "provider_sent", "reason", "price",
                "custom_reason", "complaint", "choose_language",
                "admin_menu", "reg_city", "reg_service", "reg_info", "reg_pending",
            }
            provider_menu_steps = {"provider_main", "provider_account", "provider_contact"}
            if step in client_steps:
                handle_customer(phone, text)
            elif text == "1":
                # أي "1" من المقدم = قبول طلب (حتى لو في القائمة الرئيسية)
                quoted_oid = extract_oid_from_quoted(md)
                handle_provider_accept(phone, quoted_oid=quoted_oid)
            else:
                if step not in provider_menu_steps:
                    user_sessions[phone] = {"step": "provider_main"}
                handle_provider_menu(phone, text, registered_providers[phone])
            return jsonify({"status": "ok"}), 200

        # ── عميل عادي ──
        handle_customer(phone, text)

    except Exception as e:
        print(f"Webhook error: {e}")
    return jsonify({"status": "ok"}), 200


# ==========================================
# Export Excel
# ==========================================
@app.route("/export", methods=["GET"])
def export_data():
    if request.args.get("key", "") != EXPORT_SECRET:
        return "غير مصرح ❌", 403
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()

        def header(ws, headers, color):
            for col, h in enumerate(headers, 1):
                c = ws.cell(row=1, column=col, value=h)
                c.font = Font(bold=True, color="FFFFFF")
                c.fill = PatternFill("solid", fgColor=color)
                c.alignment = Alignment(horizontal="center")
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = 20

        ws1 = wb.active
        ws1.title = "مقدمو الخدمة"
        header(ws1, ["الرقم", "الاسم", "المدينة", "التخصص", "الحالة", "الاشتراك", "التسجيل"], "1a7a4a")
        for row, (p, d) in enumerate(registered_providers.items(), 2):
            ws1.cell(row=row, column=1, value=p)
            ws1.cell(row=row, column=2, value=d.get("name", ""))
            ws1.cell(row=row, column=3, value=d.get("city", ""))
            ws1.cell(row=row, column=4, value=d.get("specialty", ""))
            ws1.cell(row=row, column=5, value="نشط" if d.get("status") == "active" else "موقوف")
            ws1.cell(row=row, column=6, value=d.get("expiry", "غير محدد"))
            ws1.cell(row=row, column=7, value=d.get("registered", ""))

        ws2 = wb.create_sheet("العملاء")
        header(ws2, ["الرقم", "الحالة"], "1a4a7a")
        for row, p in enumerate(registered_clients, 2):
            ws2.cell(row=row, column=1, value=p)
            ws2.cell(row=row, column=2, value="محظور" if p in blocked_users and time.time() < blocked_users.get(p, 0) else "نشط")

        ws3 = wb.create_sheet("الطلبات")
        header(ws3, ["رقم الطلب", "العميل", "المدينة", "الخدمة", "الوصف", "الحالة", "المحاولات", "التاريخ"], "7a4a1a")
        for row, (oid, od) in enumerate(pending_orders.items(), 2):
            ws3.cell(row=row, column=1, value=oid)
            ws3.cell(row=row, column=2, value=od.get("phone", ""))
            ws3.cell(row=row, column=3, value=od.get("city", ""))
            ws3.cell(row=row, column=4, value=od.get("service", ""))
            ws3.cell(row=row, column=5, value=od.get("description", ""))
            ws3.cell(row=row, column=6, value="مكتمل" if od.get("taken") else "معلق")
            ws3.cell(row=row, column=7, value=od.get("attempts", 1))
            ws3.cell(row=row, column=8, value=od.get("created", ""))

        ws4 = wb.create_sheet("سجل العمليات")
        header(ws4, ["الوقت", "النوع", "الرقم", "رقم الطلب", "التفاصيل"], "4a1a7a")
        ws4.column_dimensions["E"].width = 30
        for row, e in enumerate(reversed(activity_log), 2):
            ws4.cell(row=row, column=1, value=e.get("time", ""))
            ws4.cell(row=row, column=2, value=e.get("type", ""))
            ws4.cell(row=row, column=3, value=e.get("phone", ""))
            ws4.cell(row=row, column=4, value=e.get("order_id", ""))
            ws4.cell(row=row, column=5, value=e.get("details", ""))

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"mudhakkira_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        return send_file(output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True, download_name=filename)
    except ImportError:
        return "openpyxl غير مثبت", 500
    except Exception as e:
        return f"خطأ: {e}", 500


# ==========================================
# Broadcast API
# ==========================================
@app.route("/broadcast", methods=["POST"])
def broadcast_api():
    data = request.get_json()
    if not data or data.get("key") != EXPORT_SECRET:
        return jsonify({"error": "غير مصرح"}), 403
    message = data.get("message", "").strip()
    city    = data.get("city", "")
    service = data.get("service", "")
    delay   = float(data.get("delay", 3))
    if not message:
        return jsonify({"error": "الرسالة فارغة"}), 400
    targets = [
        p for p, d in registered_providers.items()
        if d.get("status") == "active"
        and (not city or d.get("city") == city)
        and (not service or d.get("specialty") == service)
    ]
    if not targets:
        return jsonify({"error": "لا يوجد مقدمون مطابقون"}), 404
    def do_broadcast():
        sent = failed = 0
        for p in targets:
            if send_msg(p, message):
                sent += 1
            else:
                failed += 1
            time.sleep(delay)
        log_event("بث_خارجي", "api", f"أُرسل: {sent} | فشل: {failed}")
        send_group(ADMIN_GROUP, f"✅ انتهى الإرسال\nأُرسل: {sent}\nفشل: {failed}")
    t = threading.Thread(target=do_broadcast)
    t.daemon = True
    t.start()
    return jsonify({
        "status": "بدأ الإرسال",
        "count": len(targets),
        "eta_minutes": int(len(targets) * delay / 60),
    }), 200


@app.route("/", methods=["GET"])
def home():
    return "مذكرة سلمان ✅", 200


# ==========================================
# التشغيل
# ==========================================
load_data()
print(f"🚀 مذكرة سلمان — Instance: {INSTANCE_ID}")

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
